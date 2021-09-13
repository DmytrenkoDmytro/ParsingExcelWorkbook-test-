#---------------Setup modules----------
import sys
import subprocess
import pkg_resources

required = {'numpy','pandas','progress','openpyxl'}
installed = {pkg.key for pkg in pkg_resources.working_set}
missing = required - installed


if missing:
    # implement pip as a subprocess:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',*missing])
#-------------------  End setup modules -------------

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook


# Load in the workbook
wb_src = load_workbook('./td_src.xlsx')
wb_upr = load_workbook('./d_upr.xlsx')
wb_vid = load_workbook('./d_vid.xlsx')
wb_ase = load_workbook('./d_ase.xlsx')
wb_result = load_workbook('./result.xlsx')


# Get a sheet by name 
sheetVP = wb_src['ВП, загвідділи']
sheetKKS = wb_src['ККС ІІІ']
sheetKGS = wb_src['КГС']
sheetKCS = wb_src['КЦС']
sheetKAS = wb_src['КАС']
sheetUPR = wb_upr['UPR']
sheetVID = wb_vid['VID']
sheetASE = wb_ase['ASE']
sheetRESULT = wb_result['RESULT']

#  BEGIN values
maxvp = sheetVP.max_row+10
maxkks = sheetKKS.max_row+10
maxkcs = sheetKCS.max_row+10
maxkas = sheetKAS.max_row+10
maxkgs = sheetKGS.max_row+10

maxvid = sheetVID.max_row
maxupr = sheetUPR.max_row
maxase = sheetASE.max_row

maxresult = maxvp+maxkks+maxkcs+maxkas+maxkgs
t=1



# progressbarInit
import time
from progress.bar import IncrementalBar

#Clearing result file
print('Clearing result file...')

bar = IncrementalBar('Progress:', max = maxresult-1)
for i in range(2,maxresult):
    
    for j in range(1,13):
        sheetRESULT.cell(row=i, column=j).value = ''
    bar.next()
    #time.sleep(1)
wb_result.save('./result.xlsx')
bar.finish()
print('Clear.')


# LOGIC Function
def Tree_f(Val_List,Val_Dict,result):
    
    LOG_result_f = Val_List == Val_Dict and Val_List != None
    if LOG_result_f: 
       result = Val_List
       
    return result







#-----------------------PARSSING VP
vid = '----'
upr = '----'
ase = '----'
firstline = 6
bar = IncrementalBar('VPZ parsing:', max = maxvp-firstline)
for i in range(firstline, maxvp):
    bar.next()
    #time.sleep(1)
    
    ValVP = sheetVP.cell(row=i, column=1).value
    v=0 
    while v<=maxvid and ValVP!=vid:
          v=v+1
          ValVID = sheetVID.cell(row=v, column=1).value
          vid = Tree_f(ValVP,ValVID,vid)
          u=0
    while u<=maxupr and ValVP!=upr:
          u=u+1
          ValUPR = sheetUPR.cell(row=u, column=1).value
          upr = Tree_f(ValVP,ValUPR,upr)
          a=0
    while a<=maxase and ValVP!=ase:
          a=a+1
          ValASE = sheetASE.cell(row=a, column=1).value
          ase = Tree_f(ValVP,ValASE,ase)
        
    
    LOG_result = ValVP != vid and ValVP != upr and ValVP != None and ValVP != ase
    if LOG_result:
       
       t=t+1 
       sheetRESULT.cell(row=t, column=1).value = 'ВП та загальні відділи'
       sheetRESULT.cell(row=t, column=2).value = ase
       sheetRESULT.cell(row=t, column=3).value = upr
       sheetRESULT.cell(row=t, column=4).value = vid
       sheetRESULT.cell(row=t, column=5).value = ValVP
       sheetRESULT.cell(row=t, column=6).value = sheetVP.cell(row=i, column=2).value
       sheetRESULT.cell(row=t, column=7).value = sheetVP.cell(row=i, column=3).value
       sheetRESULT.cell(row=t, column=8).value = sheetVP.cell(row=i, column=4).value
       sheetRESULT.cell(row=t, column=9).value = sheetVP.cell(row=i, column=5).value
       sheetRESULT.cell(row=t, column=10).value = sheetVP.cell(row=i, column=6).value
       sheetRESULT.cell(row=t, column=11).value = sheetVP.cell(row=i, column=7).value
       sheetRESULT.cell(row=t, column=12).value = sheetVP.cell(row=i, column=8).value
       
bar.finish()

#SAVE RESULT
print('Save...')
wb_result.save('./result.xlsx')
print('Saved.')

#--------------------------PARSSING KKS
vid = '----'
upr = '----'
ase = '----'
firstline = 4
bar = IncrementalBar('KKS parsing:', max = maxkks-firstline)
for i in range(firstline, maxkks):
    bar.next()
    #time.sleep(1)
    
    ValKKS = sheetKKS.cell(row=i, column=1).value
     
    v=0 
    while v<=maxvid and ValKKS!=vid:
          v=v+1
          ValVID = sheetVID.cell(row=v, column=1).value
          vid = Tree_f(ValKKS,ValVID,vid)
          u=0
    while u<=maxupr and ValKKS!=upr:
          u=u+1
          ValUPR = sheetUPR.cell(row=u, column=1).value
          upr = Tree_f(ValKKS,ValUPR,upr)
          a=0
    while a<=maxase and ValKKS!=ase:
          a=a+1
          ValASE = sheetASE.cell(row=a, column=1).value
          ase = Tree_f(ValKKS,ValASE,ase)
        
    
    LOG_result = ValKKS != vid and ValKKS != upr and ValKKS != None and ValKKS != ase
    if LOG_result:
       
       t=t+1 
       sheetRESULT.cell(row=t, column=1).value = 'КАСАЦІЙНИЙ КРИМІНАЛЬНИЙ СУД'
       sheetRESULT.cell(row=t, column=2).value = ase
       sheetRESULT.cell(row=t, column=3).value = upr
       sheetRESULT.cell(row=t, column=4).value = vid
       sheetRESULT.cell(row=t, column=5).value = ValKKS
       sheetRESULT.cell(row=t, column=6).value = sheetKKS.cell(row=i, column=2).value
       sheetRESULT.cell(row=t, column=7).value = sheetKKS.cell(row=i, column=3).value
       sheetRESULT.cell(row=t, column=8).value = sheetKKS.cell(row=i, column=4).value
       sheetRESULT.cell(row=t, column=9).value = sheetKKS.cell(row=i, column=5).value
       sheetRESULT.cell(row=t, column=10).value = sheetKKS.cell(row=i, column=6).value
       sheetRESULT.cell(row=t, column=11).value = sheetKKS.cell(row=i, column=7).value
       sheetRESULT.cell(row=t, column=12).value = sheetKKS.cell(row=i, column=8).value
       
bar.finish()

#SAVE RESULT
print('Save...')
wb_result.save('./result.xlsx')
print('Saved.')

#-------------------------PARSSING KGS
vid = '----'
upr = '----'
ase = '----'
firstline = 7
bar = IncrementalBar('KGS parsing:', max = maxkgs-firstline)
for i in range(firstline, maxkgs):
    bar.next()
    #time.sleep(1)
    
    ValKGS = sheetKGS.cell(row=i, column=1).value
     
    v=0 
    while v<=maxvid and ValKGS!=vid:
          v=v+1
          ValVID = sheetVID.cell(row=v, column=1).value
          vid = Tree_f(ValKGS,ValVID,vid)
          u=0
    while u<=maxupr and ValKGS!=upr:
          u=u+1
          ValUPR = sheetUPR.cell(row=u, column=1).value
          upr = Tree_f(ValKGS,ValUPR,upr)
          a=0
    while a<=maxase and ValKGS!=ase:
          a=a+1
          ValASE = sheetASE.cell(row=a, column=1).value
          ase = Tree_f(ValKGS,ValASE,ase)
        
    
    LOG_result = ValKGS != vid and ValKGS != upr and ValKGS != None and ValKGS != ase
    if LOG_result:
       
       t=t+1 
       sheetRESULT.cell(row=t, column=1).value = 'Касаційний господарський суд'
       sheetRESULT.cell(row=t, column=2).value = ase
       sheetRESULT.cell(row=t, column=3).value = upr
       sheetRESULT.cell(row=t, column=4).value = vid
       sheetRESULT.cell(row=t, column=5).value = ValKGS
       sheetRESULT.cell(row=t, column=6).value = sheetKGS.cell(row=i, column=2).value
       sheetRESULT.cell(row=t, column=7).value = sheetKGS.cell(row=i, column=3).value
       sheetRESULT.cell(row=t, column=8).value = sheetKGS.cell(row=i, column=4).value
       sheetRESULT.cell(row=t, column=9).value = sheetKGS.cell(row=i, column=5).value
       sheetRESULT.cell(row=t, column=10).value = sheetKGS.cell(row=i, column=6).value
       sheetRESULT.cell(row=t, column=11).value = sheetKGS.cell(row=i, column=7).value
       sheetRESULT.cell(row=t, column=12).value = sheetKGS.cell(row=i, column=8).value
       
bar.finish()

#SAVE RESULT
print('Save...')
wb_result.save('./result.xlsx')
print('Saved.')



#----------------------------PARSSING KCS
vid = '----'
upr = '----'
ase = '----'
firstline = 6
bar = IncrementalBar('KCS parsing:', max = maxkcs-firstline)
for i in range(firstline, maxkcs):
    bar.next()
    #time.sleep(1)
    
    ValKCS = sheetKCS.cell(row=i, column=1).value
     
    v=0 
    while v<=maxvid and ValKCS!=vid:
          v=v+1
          ValVID = sheetVID.cell(row=v, column=1).value
          vid = Tree_f(ValKCS,ValVID,vid)
          u=0
    while u<=maxupr and ValKCS!=upr:
          u=u+1
          ValUPR = sheetUPR.cell(row=u, column=1).value
          upr = Tree_f(ValKCS,ValUPR,upr)
          a=0
    while a<=maxase and ValKCS!=ase:
          a=a+1
          ValASE = sheetASE.cell(row=a, column=1).value
          ase = Tree_f(ValKCS,ValASE,ase)
        
    
    LOG_result = ValKCS != vid and ValKCS != upr and ValKCS != None and ValKCS != ase
    if LOG_result:
       
       t=t+1 
       sheetRESULT.cell(row=t, column=1).value = 'Касаційний цивільний суд'
       sheetRESULT.cell(row=t, column=2).value = ase
       sheetRESULT.cell(row=t, column=3).value = upr
       sheetRESULT.cell(row=t, column=4).value = vid
       sheetRESULT.cell(row=t, column=5).value = ValKCS
       sheetRESULT.cell(row=t, column=6).value = sheetKCS.cell(row=i, column=2).value
       sheetRESULT.cell(row=t, column=7).value = sheetKCS.cell(row=i, column=3).value
       sheetRESULT.cell(row=t, column=8).value = sheetKCS.cell(row=i, column=4).value
       sheetRESULT.cell(row=t, column=9).value = sheetKCS.cell(row=i, column=5).value
       sheetRESULT.cell(row=t, column=10).value = sheetKCS.cell(row=i, column=6).value
       sheetRESULT.cell(row=t, column=11).value = sheetKCS.cell(row=i, column=7).value
       sheetRESULT.cell(row=t, column=12).value = sheetKCS.cell(row=i, column=8).value
       
bar.finish()

#SAVE RESULT
print('Save...')
wb_result.save('./result.xlsx')
print('Saved.')

#------------------------------PARSSING KAS
vid = '----'
upr = '----'
ase = '----'
firstline = 6
bar = IncrementalBar('KAS parsing:', max = maxkas-firstline)
for i in range(firstline, maxkas):
    bar.next()
    #time.sleep(1)
    
    ValKAS = sheetKAS.cell(row=i, column=1).value
     
    v=0 
    while v<=maxvid and ValKAS!=vid:
          v=v+1
          ValVID = sheetVID.cell(row=v, column=1).value
          vid = Tree_f(ValKAS,ValVID,vid)
          u=0
    while u<=maxupr and ValKAS!=upr:
          u=u+1
          ValUPR = sheetUPR.cell(row=u, column=1).value
          upr = Tree_f(ValKAS,ValUPR,upr)
          a=0
    while a<=maxase and ValKAS!=ase:
          a=a+1
          ValASE = sheetASE.cell(row=a, column=1).value
          ase = Tree_f(ValKAS,ValASE,ase)
        
    
    LOG_result = ValKAS != vid and ValKAS != upr and ValKAS != None and ValKAS != ase
    if LOG_result:
       
       t=t+1 
       sheetRESULT.cell(row=t, column=1).value = 'Касаційний адміністративний суд'
       sheetRESULT.cell(row=t, column=2).value = ase
       sheetRESULT.cell(row=t, column=3).value = upr
       sheetRESULT.cell(row=t, column=4).value = vid
       sheetRESULT.cell(row=t, column=5).value = ValKAS
       sheetRESULT.cell(row=t, column=6).value = sheetKAS.cell(row=i, column=2).value
       sheetRESULT.cell(row=t, column=7).value = sheetKAS.cell(row=i, column=3).value
       sheetRESULT.cell(row=t, column=8).value = sheetKAS.cell(row=i, column=4).value
       sheetRESULT.cell(row=t, column=9).value = sheetKAS.cell(row=i, column=5).value
       sheetRESULT.cell(row=t, column=10).value = sheetKAS.cell(row=i, column=6).value
       sheetRESULT.cell(row=t, column=11).value = sheetKAS.cell(row=i, column=7).value
       sheetRESULT.cell(row=t, column=12).value = sheetKAS.cell(row=i, column=8).value
       
bar.finish()


#SAVE RESULT
print('Save...')
wb_result.save('./result.xlsx')
print('Saved!')
print('------------------------------')
varend=input ('          Press Enter...') 
